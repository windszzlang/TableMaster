import re
import openpyxl
import math
import numpy as np
import pandas as pd

# xlsx is different from xls in terms of the internal structure of the file
# openpyxl for xlsx
# xlrd for xls
# pandas can only read data not including formatting, styles, etc.



def get_sheet_names(file_path):
    workbook = openpyxl.load_workbook(filename=file_path)
    sheet_names = workbook.sheetnames
    return sheet_names


def read_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(filename=file_path)
    sheet = workbook[sheet_name]

    # print(sheet)
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    
    return data


def read_csv(file_path):
    df = pd.read_csv(file_path, on_bad_lines='warn')
    data = df.values.tolist()
    return data


def add_address_to_table(table, start_row, start_col):
    table = table.strip().split('\n')
    
    for i, row in enumerate(table):
        cells = [cell.strip() for cell in row.split('|')]
        
        for j, cell in enumerate(cells):
            cells[j] = f"{index_to_column(j + start_col)}{i + start_row + 1},{cell}"
        table[i] = '| ' +  ' | '.join(cells) + ' |'
    return '\n'.join(table)


# zero-base index
def index_to_column(number):
    result = ""
    while number >= 0:
        result = chr(number % 26 + 65) + result
        number = number // 26 - 1 
    return result


def column_to_index(column):
    result = 0
    for char in column.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def cell_to_index(cell):
    """
    Convert an Excel-style cell (e.g., 'A1', 'AA10') to (row, column) indices.
    """
    match = re.match(r"([A-Z]+)(\d+)", cell)
    if match:
        col_str, row_str = match.groups()
        col = column_to_index(col_str)
        row = int(row_str) - 1  # Excel rows are 1-indexed, Python uses 0-indexing
        return row, col
    else:
        raise ValueError(f"Invalid cell format: {cell}")


def compress_table(table:list[list[str]], max_char=100):
    compressed_table = [
        [cell[:max_char] for cell in row]
        for row in table
    ]
    return compressed_table


def format_table(table:list[list[str]], with_address=False) -> str:
    """
    Format the table into a string
    """
    table_str = ''
    for i, row in enumerate(table):
        row_str = '|'
        for j, cell in enumerate(row):
            if with_address:
                row_str += f'{index_to_column(j)}{i+1},{cell}|'
            else:
                row_str += f'{cell}|'
        table_str += row_str +'\n'
    return table_str

def remove_none_in_table(table, replace_with='None'):
    for i, row in enumerate(table):
        for j, cell in enumerate(row):
            if cell == None or (isinstance(cell, float) and math.isnan(cell)):
                table[i][j] = replace_with
    return table


def parse_range(range_str):
    """
    Parse a cell range (e.g., 'A1:AA10') into a list of (row, column) tuples.
    """
    start_cell, end_cell = range_str.split(':')
    start_row, start_col = cell_to_index(start_cell)
    end_row, end_col = cell_to_index(end_cell)
    
    cells = []
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cells.append((row, col))
    
    return cells


def split_table_by_rows(table, row_number):
    '''
    |a1,1|b1,2|c1,3|
    |a2,1|b2,2|c2,3|
    '''
    first_col = table[1]
    rows = table.split(f'\n|{first_col}')
    rows = [rows[0]] + [f"|{first_col}" + row for row in rows[1:]]

    split_tables = [rows[i:i + row_number] for i in range(0, len(rows), row_number)]

    split_tables = ['\n'.join(part) for part in split_tables]
    
    return split_tables


def truncate_table_by_rows(table, row_number):
    split_tables = split_table_by_rows(table, row_number)

    return split_tables[0]


def merge_regions(regions):
    """
    Merge regions as much as possible and return a list of merged and unmerged regions.
    """
    def are_regions_mergeable(region1, region2):
        """
        Check if two regions are adjacent or overlapping and can be merged.
        """
        r1_cells = set(parse_range(region1))
        r2_cells = set(parse_range(region2))
        
        # Check if there is any overlap or adjacency between the two regions
        if r1_cells & r2_cells:  # If there is any common cell, they overlap
            return True
        
        # Check if any cell in region1 is adjacent to region2 (contiguous horizontally or vertically)
        for r1_row, r1_col in r1_cells:
            for r2_row, r2_col in r2_cells:
                if abs(r1_row - r2_row) <= 1 and r1_col == r2_col:  # Vertical adjacency
                    return True
                if abs(r1_col - r2_col) <= 1 and r1_row == r2_row:  # Horizontal adjacency
                    return True
        
        return False


    def merge_two_regions(region1, region2):
        """
        Merge two regions into one larger region.
        """
        r1_cells = parse_range(region1)
        r2_cells = parse_range(region2)
        
        all_cells = r1_cells + r2_cells
        all_rows = [cell[0] for cell in all_cells]
        all_cols = [cell[1] for cell in all_cells]
        
        min_row, max_row = min(all_rows), max(all_rows)
        min_col, max_col = min(all_cols), max(all_cols)
        
        start_cell = index_to_column(min_col) + str(min_row + 1)
        end_cell = index_to_column(max_col) + str(max_row + 1)
        
        return f"{start_cell}:{end_cell}"

    merged_regions = []
    
    while regions:
        region = regions.pop(0)
        merged = False
        
        # Try to merge the current region with any of the already merged regions
        for i, merged_region in enumerate(merged_regions):
            if are_regions_mergeable(region, merged_region):
                merged_regions[i] = merge_two_regions(region, merged_region)
                merged = True
                break
        
        # If it cannot be merged, add it to the merged regions as a separate one
        if not merged:
            merged_regions.append(region)
    
    return merged_regions



def execute_python_code(code: str, result_var_name: str = 'answer'):
    """
    Executes a Python code string and returns the result of the last expression.
    """
    # Create a local dictionary to store the result of the execution
    local_vars = {}
    try:
        # Execute the code and capture any variables in local_vars
        exec(code, {}, local_vars)
        
        # Try to return the result of the last expression, if available
        result = local_vars.get(result_var_name, None)
        if result is None:
            return 'Error: No explicit result found.'
        return result
    except Exception as e:
        # Handle and return any errors that occur during execution
        return f"Error: {str(e)}"



if __name__ == "__main__":
    # test_data ={'table_array': [['a', 'b', 'c'], ['d', 'e', 'f'], ['g', 'h', 'i']]}
    # print(format_table(test_data['table_array']))

    print(index_to_column(0))
